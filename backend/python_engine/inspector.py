import sys
import openpyxl

def main():
    # 1. Windows Safe Printing
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')

    # 2. Get File Path
    if len(sys.argv) < 2:
        print("Usage: python inspector.py <path_to_excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    print(f"--- INSPECTING: {file_path} ---")

    try:
        # 3. Load Workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 4. Print Sheet Names
        print(f"\n[SHEET NAMES FOUND]:")
        for name in wb.sheetnames:
            print(f" - '{name}'")

        # 5. Scan the First Sheet (usually Basic Info)
        first_sheet = wb.sheetnames[0]
        ws = wb[first_sheet]
        print(f"\n[SCANNING FIRST 15 ROWS OF '{first_sheet}']:")
        print("Row | Col A | Col B | Col C | Col D |")
        print("-" * 40)

        for row in ws.iter_rows(min_row=1, max_row=15, max_col=5):
            row_num = row[0].row
            # Grab values for first 5 columns (A, B, C, D, E)
            vals = [str(cell.value).strip() if cell.value else "" for cell in row]
            
            # Print cleanly
            print(f" {row_num:2} | {vals[0]:15} | {vals[1]:15} | {vals[2]:15} | {vals[3]:15} |")

    except Exception as e:
        print(f"ERROR: {str(e)}")

if __name__ == "__main__":
    main()