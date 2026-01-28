import sys
import json
import openpyxl
import io

# ==========================================
# CONSOLE ENCODING FIX
# ==========================================
try:
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

def safe_write(ws, coord, value):
    """
    Handles Merged Cells: Finds the master cell (top-left) 
    of a merged range to ensure data is actually visible.
    """
    val_str = str(value)
    # Check if the coordinate is part of any merged range
    for range_ in ws.merged_cells.ranges:
        if coord in range_:
            top_left = range_.start_cell
            ws[top_left.coordinate] = val_str
            return 
    # Normal write if not merged
    ws[coord] = val_str

def find_sheet(wb, keyword):
    """
    Case-insensitive search for sheet names to make the 
    system compatible with different KRA template versions.
    """
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return wb[name]
    return None

def main():
    if len(sys.argv) < 4:
        print("Error: Missing arguments (Template, Data, Output)")
        sys.exit(1)

    template_path, json_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3]

    try:
        print(f"Opening template: {template_path}")
        # keep_vba=True is required to preserve the 'Validate' button logic
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        
        # UX FIX: Accessing the vba_archive forces openpyxl to re-link 
        # macro objects (buttons) which helps prevent them from disappearing.
        if wb.vba_archive:
            _ = wb.vba_archive 

        with open(json_path, "r", encoding='utf-8') as f:
            data = json.load(f)
        
        instructions = data.get("instructions", [])
        print(f"Processing {len(instructions)} autofill instructions...")

        for inst in instructions:
            ws = find_sheet(wb, inst['sheet_keyword'])
            if ws:
                # KRA sheets are protected; unprotect in memory to allow writing
                ws.protection.sheet = False 
                safe_write(ws, inst['cell'], inst['value'])
            else:
                print(f"Warning: Sheet matching '{inst['sheet_keyword']}' not found.")
        
        print(f"Saving finalized return to: {output_path}")
        wb.save(output_path)
        print("SUCCESS: KRA Return generated successfully.")

    except Exception as e:
        print(f"CRITICAL ERROR: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()